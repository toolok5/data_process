import os
import pandas as pd
from tkinter import Tk, simpledialog
from tkinter.filedialog import askopenfilenames
import random
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def parse_column_range(column_range):
    try:
        if '-' in column_range:
            start, end = map(int, column_range.split('-'))
            return list(range(start - 1, end))  # 0 基索引
        else:
            return [int(column_range) - 1]  # 单列情况
    except ValueError:
        raise ValueError("无效的列范围输入，请输入数字或范围（例如：1 或 1-3）。")

def get_row_count(file_path, columns):
    try:
        if file_path.endswith(".csv"):
            encodings = ['gbk', 'utf-8', 'latin1']
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, 
                                   encoding=encoding, 
                                   usecols=columns, 
                                   low_memory=False,
                                   dtype=str)
                    row_count = len(df.dropna(how='all'))
                    return [{"sheet_name": "CSV", "row_count": row_count}]
                except Exception as e:
                    print(f"尝试使用编码 {encoding} 读取 CSV 文件失败：{e}")
            raise ValueError("无法读取 CSV 文件，可能编码不支持")

        elif file_path.endswith(('.xls', '.xlsx')):
            try:
                wb = load_workbook(filename=file_path, read_only=True)
                results = []
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    row_count = 0
                    is_first_row = True  # 添加标志来识别表头行
                    
                    # 使用iter_rows只读取指定列
                    for row in ws.iter_rows(min_col=min(columns)+1, 
                                          max_col=max(columns)+1,
                                          values_only=True):
                        if any(cell is not None and str(cell).strip() for cell in row):
                            if not is_first_row:  # 如果不是第一行，才计数
                                row_count += 1
                            is_first_row = False  # 处理完第一行后，将标志设为False
                            
                    results.append({
                        "sheet_name": sheet_name,
                        "row_count": row_count  # 这里的row_count已经不包含表头行
                    })
                
                wb.close()
                return results
                
            except Exception as e:
                print(f"读取Excel文件出错：{e}")
                return [{"sheet_name": "错误", "row_count": "读取失败"}]
        else:
            raise ValueError("不支持的文件格式")
    except Exception as e:
        print(f"读取文件 {file_path} 出错：{e}")
        return [{"sheet_name": "错误", "row_count": "读取失败"}]

def append_to_existing_file(existing_path, new_data):
    try:
        existing_df = pd.read_excel(existing_path, engine="openpyxl")
        if len(existing_df) != len(new_data):
            max_len = max(len(existing_df), len(new_data))
            existing_df = existing_df.reindex(range(max_len))
            new_data = new_data.reindex(range(max_len))
        for column in new_data.columns:
            new_col_name = column
            while new_col_name in existing_df.columns:
                new_col_name = f"{column}_{random.randint(1, 1000)}"
            existing_df[new_col_name] = new_data[column].values
    except Exception as e:
        print(f"读取或合并现有文件出错：{e}")
        existing_df = new_data
    existing_df.to_excel(existing_path, index=False)
    print(f"结果已追加保存到：{existing_path}")

def count_file_rows():
    root = Tk()
    root.withdraw()

    column_range = simpledialog.askstring("输入列范围", 
                                        "请输入需要读取的列范围（默认: 1，例如: 1 或 1-3）：", 
                                        initialvalue="1",
                                        parent=root)
    if not column_range:
        print("未输入列范围，操作已取消。")
        return

    try:
        columns = parse_column_range(column_range)
    except ValueError as e:
        print(e)
        return

    file_paths = askopenfilenames(
        filetypes=[("Excel or CSV Files", "*.xls *.xlsx *.csv")], 
        title="选择文件",
        parent=root
    )
    
    if not file_paths:
        print("未选择任何文件")
        return

    stats = []
    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        print(f"正在处理文件: {file_name}")  # 添加进度提示
        
        sheet_stats = get_row_count(file_path, columns)
        for sheet_info in sheet_stats:
            stats.append({
                "文件名": file_name,
                "工作表名": sheet_info["sheet_name"],
                "行数": sheet_info["row_count"]
            })
            print(f"文件: {file_name}, 工作表: {sheet_info['sheet_name']}, 行数: {sheet_info['row_count']}")

    stats_df = pd.DataFrame(stats)
    save_dir = r"C:\\excel"
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "行数统计结果.xlsx")

    if os.path.exists(save_path):
        append_to_existing_file(save_path, stats_df)
    else:
        stats_df.to_excel(save_path, index=False)
        print(f"结果已保存到: {save_path}")

def main():
    count_file_rows()

if __name__ == "__main__":
    main()
